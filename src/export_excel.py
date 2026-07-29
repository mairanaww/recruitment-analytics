import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule
import os
import sys

# Ensure src modules can be imported
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from src.etl import get_engine

def export_to_excel():
    print("Connecting to database...")
    engine = get_engine()
    
    views = {
        'Time to Fill': 'v_time_to_fill',
        'Funnel Conversion': 'v_funnel_conversion',
        'Offer Acceptance': 'v_offer_acceptance',
        'Headcount vs Plan': 'v_headcount_vs_plan'
    }
    
    wb = Workbook()
    wb.remove(wb.active) # Remove default sheet
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    print("Exporting views...")
    for sheet_name, view_name in views.items():
        ws = wb.create_sheet(title=sheet_name)
        
        try:
            df = pd.read_sql_query(f"SELECT * FROM {view_name}", engine)
        except Exception as e:
            print(f"Error reading view {view_name}: {e}")
            continue
            
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                    
        # Apply some conditional formatting as an example
        if sheet_name == 'Offer Acceptance':
            # Conditional formatting for acceptance_rate_pct (column D)
            rule = ColorScaleRule(start_type='min', start_color='F8696B',
                                  mid_type='percentile', mid_value=50, mid_color='FFEB84',
                                  end_type='max', end_color='63BE7B')
            ws.conditional_formatting.add(f'D2:D{len(df)+1}', rule) 
            
    os.makedirs('reports', exist_ok=True)
    report_path = 'reports/weekly_recruitment_summary.xlsx'
    wb.save(report_path)
    print(f"Report saved to {report_path}")

if __name__ == '__main__':
    export_to_excel()
